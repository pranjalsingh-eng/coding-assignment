"""
Q3: Characters of Ice and Fire
Tasks:
  a. Get all characters from API
  b. Find how many seasons the character was part of (via tvSeries list)
  c. Sort according to number of season appearances (descending)
  d. Add all available sorted data into an Excel file
API URL: https://anapioficeandfire.com/api/characters
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


API_URL = "https://anapioficeandfire.com/api/characters"
OUTPUT_EXCEL = "characters_output.xlsx"


def fetch_all_characters(base_url: str) -> list[dict]:
    """Fetch all characters from the API handling pagination."""
    characters = []
    page = 1
    page_size = 50

    while True:
        params = {"page": page, "pageSize": page_size}
        response = requests.get(base_url, params=params, timeout=15)
        response.raise_for_status()

        data = response.json()
        if not data:
            break

        characters.extend(data)
        print(f"  Fetched page {page} — {len(data)} characters  (total so far: {len(characters)})")

        if len(data) < page_size:
            break
        page += 1

    return characters


def parse_character(char: dict) -> dict:
    """Extract and compute relevant fields for a character."""
    name = char.get("name") or "Unknown"
    gender = char.get("gender") or "Unknown"
    culture = char.get("culture") or "Unknown"
    born = char.get("born") or "Unknown"
    died = char.get("died") or "Unknown"
    tv_series = [s for s in char.get("tvSeries", []) if s]  # filter empty strings
    season_count = len(tv_series)
    aliases = ", ".join(char.get("aliases", [])) or "None"
    titles = ", ".join(char.get("titles", [])) or "None"
    books_count = len([b for b in char.get("books", []) if b])
    povbooks_count = len([b for b in char.get("povBooks", []) if b])
    played_by = ", ".join(char.get("playedBy", [])) or "Unknown"

    return {
        "Name": name,
        "Gender": gender,
        "Culture": culture,
        "Born": born,
        "Died": died,
        "Season Count": season_count,
        "TV Seasons": ", ".join(tv_series) if tv_series else "None",
        "Aliases": aliases,
        "Titles": titles,
        "Books Count": books_count,
        "POV Books Count": povbooks_count,
        "Played By": played_by,
    }


def write_excel(characters_data: list[dict], filepath: str) -> None:
    """Write character data to an Excel file with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Characters"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Alternate row fill
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    headers = list(characters_data[0].keys())
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for i, row_data in enumerate(characters_data, start=2):
        ws.append(list(row_data.values()))
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = alt_fill
        for cell in ws[i]:
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Set column widths
    col_widths = [30, 10, 15, 20, 20, 14, 50, 30, 35, 12, 14, 25]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(filepath)


def main():
    print("=" * 60)
    print("Q3: Characters of Ice and Fire")
    print("=" * 60)

    # a. Fetch all characters
    print("\n[a] Fetching all characters from API...")
    raw_characters = fetch_all_characters(API_URL)
    print(f"    Total characters fetched: {len(raw_characters)}")

    # b & c. Parse and sort by season count (descending)
    print("\n[b] Computing season appearances for each character...")
    parsed = [parse_character(c) for c in raw_characters]

    print("\n[c] Sorting by number of season appearances (descending)...")
    parsed.sort(key=lambda x: x["Season Count"], reverse=True)

    # Preview top 10
    print("\n    Top 10 characters by season appearances:")
    for char in parsed[:10]:
        print(f"      {char['Name']:<30} Seasons: {char['Season Count']}")

    # d. Write to Excel
    print(f"\n[d] Writing sorted data to '{OUTPUT_EXCEL}'...")
    write_excel(parsed, OUTPUT_EXCEL)
    print(f"    File written: {OUTPUT_EXCEL}")
    print(f"\nDone! {len(parsed)} characters saved to Excel.")
    return parsed


if __name__ == "__main__":
    main()